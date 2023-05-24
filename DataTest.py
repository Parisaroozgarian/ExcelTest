import openpyxl

wb1=openpyxl.load_workbook("C:\\Users\\AUSU\\PycharmProjects\\pythonProject2\\DataTest\\DataTest.xlsx")
sheets =wb1.sheetnames
print(wb1.active.title)
sh1=wb1['Data1']
data=sh1['A2'].value
data2=wb1['Data1']['A3'].value
print(sh1.cell(1,2).value)
print(sh1.cell(2,3).value)

